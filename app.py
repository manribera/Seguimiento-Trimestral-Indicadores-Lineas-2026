# app.py
# -*- coding: utf-8 -*-

import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

st.set_page_config(page_title="Seguimiento de líneas de acción", layout="wide")

# =========================================================
# SESSION STATE
# =========================================================
if "lineas_guardadas" not in st.session_state:
    st.session_state["lineas_guardadas"] = {}

if "archivo_nombre" not in st.session_state:
    st.session_state["archivo_nombre"] = ""

if "pdf_final" not in st.session_state:
    st.session_state["pdf_final"] = None


# =========================================================
# UTILIDADES DE TEXTO
# =========================================================
def norm_text(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_meaningful(value) -> bool:
    return value not in (None, "")


# =========================================================
# UTILIDADES DE EXCEL
# =========================================================
def get_effective_cell_value(ws, row, col):
    """
    Devuelve el valor real de una celda, incluyendo celdas combinadas.
    """
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def row_values(ws, row, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    return [get_effective_cell_value(ws, row, c) for c in range(1, max_col + 1)]


def row_text(ws, row, max_col=None):
    vals = row_values(ws, row, max_col)
    return " | ".join("" if v is None else str(v) for v in vals)


def get_right_value(ws, row, col, max_steps=8):
    for c in range(col + 1, min(ws.max_column, col + max_steps) + 1):
        val = get_effective_cell_value(ws, row, c)
        if is_meaningful(val):
            return val
    return ""


def get_down_value(ws, row, col, max_steps=4):
    for r in range(row + 1, min(ws.max_row, row + max_steps) + 1):
        val = get_effective_cell_value(ws, r, col)
        if is_meaningful(val):
            return val
    return ""


def get_near_value(ws, row, col):
    val = get_right_value(ws, row, col, max_steps=10)
    if is_meaningful(val):
        return val

    val = get_down_value(ws, row, col, max_steps=4)
    if is_meaningful(val):
        return val

    return ""


# =========================================================
# DETECCIÓN DE HOJA PRINCIPAL
# =========================================================
def find_best_main_sheet(wb):
    candidates = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        score = 0

        max_row = min(ws.max_row, 220)
        max_col = min(ws.max_column, 50)

        for r in range(1, max_row + 1):
            txt = norm_text(row_text(ws, r, max_col))

            if "línea de acción" in txt or "linea de accion" in txt:
                score += 8
            if "problemática" in txt or "problematica" in txt:
                score += 6
            if "líder estratégico" in txt or "lider estrategico" in txt:
                score += 6
            if "indicador" in txt:
                score += 4
            if "meta" in txt:
                score += 3
            if "avance" in txt:
                score += 3
            if "descripción" in txt or "descripcion" in txt:
                score += 3
            if "cantidad" in txt:
                score += 3
            if "observaciones" in txt or "observación" in txt:
                score += 2
            if "delegación" in txt or "delegacion" in txt:
                score += 3
            if "trimestre" in txt:
                score += 2

        candidates.append((sheet_name, score))

    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates[0][0] if candidates else wb.sheetnames[0]


# =========================================================
# DATOS GENERALES
# =========================================================
def get_delegacion(ws):
    for r in range(1, min(ws.max_row, 80) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            val = get_effective_cell_value(ws, r, c)
            if "delegación" in norm_text(val) or "delegacion" in norm_text(val):
                derecha = get_right_value(ws, r, c, max_steps=8)
                if is_meaningful(derecha):
                    return clean_text(derecha)

                abajo = get_down_value(ws, r, c, max_steps=3)
                if is_meaningful(abajo):
                    return clean_text(abajo)
    return ""


# =========================================================
# LÍNEA DE ACCIÓN
# =========================================================
def looks_like_bad_line_value(text: str) -> bool:
    t = norm_text(text)

    bad_patterns = [
        "problemática",
        "problematica",
        "problemática de linea",
        "problematica de linea",
        "líder",
        "lider",
        "delegación",
        "delegacion",
        "municipalidad",
        "trimestre",
        "indicador",
        "meta",
        "avance",
        "descripción",
        "descripcion",
        "cantidad",
        "observaciones",
        "línea de acción",
        "linea de accion",
    ]
    return any(p in t for p in bad_patterns)


def extract_line_number_from_area(ws, start_row, start_col):
    candidates = []

    for c in range(start_col + 1, min(ws.max_column, start_col + 8) + 1):
        val = get_effective_cell_value(ws, start_row, c)
        if is_meaningful(val):
            txt = clean_text(val)
            if not looks_like_bad_line_value(txt):
                candidates.append(txt)

    for r in range(start_row, min(ws.max_row, start_row + 3) + 1):
        for c in range(start_col, min(ws.max_column, start_col + 6) + 1):
            val = get_effective_cell_value(ws, r, c)
            if is_meaningful(val):
                txt = clean_text(val)
                if not looks_like_bad_line_value(txt):
                    candidates.append(txt)

    if not candidates:
        return ""

    for x in candidates:
        m = re.search(r"\d+([.\-]\d+)?", str(x))
        if m:
            return m.group(0)

    short_candidates = [x for x in candidates if len(str(x).strip()) <= 10]
    if short_candidates:
        return str(short_candidates[0]).strip()

    return str(candidates[0]).strip()


def search_value_near_keywords(ws, start_row, end_row, keywords):
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            t = norm_text(val)
            if any(k in t for k in keywords):
                cerca = get_near_value(ws, r, c)
                if is_meaningful(cerca):
                    return clean_text(cerca)
    return ""


def detect_trimester(ws, start_row, end_row):
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        txt = norm_text(row_text(ws, r))

        if "trimestre" in txt:
            for c in range(1, ws.max_column + 1):
                v = clean_text(get_effective_cell_value(ws, r, c))
                if v in ["I", "II", "III", "IV"]:
                    return v

            if " iv" in f" {txt}" or "cuarto" in txt or "4" in txt:
                return "IV"
            if " iii" in f" {txt}" or "tercer" in txt or "3" in txt:
                return "III"
            if " ii" in f" {txt}" or "segundo" in txt or "2" in txt:
                return "II"
            if " i" in f" {txt}" or "primer" in txt or "1" in txt:
                return "I"

    return ""


# =========================================================
# BLOQUES
# =========================================================
def find_line_action_starts(ws):
    starts = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = get_effective_cell_value(ws, r, c)
            t = norm_text(val)

            if "línea de acción" in t or "linea de accion" in t:
                line_value = extract_line_number_from_area(ws, r, c)
                starts.append({
                    "row": r,
                    "col": c,
                    "line_value": line_value
                })
                break

    cleaned = []
    last_row = -999

    # Más tolerancia para no perder líneas
    for item in starts:
        if item["row"] - last_row > 1:
            cleaned.append(item)
            last_row = item["row"]

    return cleaned


def detect_header_row(ws, start_row, end_row):
    best_row = None
    best_score = -1

    for r in range(start_row, min(end_row, ws.max_row) + 1):
        vals = row_values(ws, r)

        found = {
            "indicador": False,
            "meta": False,
            "avance": False,
            "descripcion": False,
            "cantidad": False,
            "observaciones": False,
        }

        for v in vals:
            t = norm_text(v)
            if "indicador" in t:
                found["indicador"] = True
            if "meta" in t:
                found["meta"] = True
            if "avance" in t:
                found["avance"] = True
            if "descripcion" in t or "descripción" in t:
                found["descripcion"] = True
            if "cantidad" in t:
                found["cantidad"] = True
            if "observ" in t:
                found["observaciones"] = True

        score = sum(found.values())

        if found["indicador"] and found["meta"] and score > best_score:
            best_score = score
            best_row = r

    return best_row


def map_headers(ws, header_row):
    header_map = {}

    for c in range(1, ws.max_column + 1):
        val = get_effective_cell_value(ws, header_row, c)
        t = norm_text(val)

        if "indicador" in t:
            header_map["Indicador"] = c
        elif "meta" in t:
            header_map["Meta (editable)"] = c
        elif "avance" in t:
            header_map["Avance"] = c
        elif "descripcion" in t or "descripción" in t:
            header_map["Descripción"] = c
        elif "cantidad" in t:
            header_map["Cantidad"] = c
        elif "observ" in t:
            header_map["Observaciones (Editable)"] = c

    return header_map


def extract_table(ws, header_row, block_end_row):
    columns = [
        "Indicador",
        "Meta (editable)",
        "Avance",
        "Descripción",
        "Cantidad",
        "Observaciones (Editable)"
    ]

    header_map = map_headers(ws, header_row)

    if "Indicador" not in header_map:
        return pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance": "",
            "Descripción": "",
            "Cantidad": "",
            "Observaciones (Editable)": ""
        }])

    data = []
    empty_count = 0

    for r in range(header_row + 1, block_end_row + 1):
        row_data = {}

        for col_name in columns:
            col_idx = header_map.get(col_name)
            row_data[col_name] = get_effective_cell_value(ws, r, col_idx) if col_idx else ""

        for k in row_data:
            if row_data[k] is None:
                row_data[k] = ""

        has_content = any(str(v).strip() != "" for v in row_data.values())

        if not has_content:
            empty_count += 1
            if empty_count >= 4:
                break
            continue

        empty_count = 0
        data.append(row_data)

    df = pd.DataFrame(data, columns=columns)

    if df.empty:
        df = pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance": "",
            "Descripción": "",
            "Cantidad": "",
            "Observaciones (Editable)": ""
        }])

    return df


def extract_blocks_from_sheet(ws):
    starts = find_line_action_starts(ws)
    delegacion = get_delegacion(ws)
    blocks = []

    if not starts:
        return blocks

    for i, start in enumerate(starts):
        start_row = start["row"]
        end_row = starts[i + 1]["row"] - 1 if i + 1 < len(starts) else ws.max_row

        linea_numero = start["line_value"]

        problematica = search_value_near_keywords(
            ws,
            start_row,
            min(start_row + 12, end_row),
            ["problemática", "problematica"]
        )

        lider = search_value_near_keywords(
            ws,
            start_row,
            min(start_row + 12, end_row),
            ["líder estratégico", "lider estrategico", "líder", "lider"]
        )

        trimestre = detect_trimester(
            ws,
            start_row,
            min(start_row + 12, end_row)
        )

        header_row = detect_header_row(
            ws,
            start_row,
            min(start_row + 30, end_row)
        )

        tabla = extract_table(ws, header_row, end_row) if header_row else pd.DataFrame([{
            "Indicador": "",
            "Meta (editable)": "",
            "Avance": "",
            "Descripción": "",
            "Cantidad": "",
            "Observaciones (Editable)": ""
        }])

        blocks.append({
            "delegacion": delegacion,
            "linea_accion": linea_numero if linea_numero else str(i + 1),
            "problematica": problematica,
            "lider": lider,
            "trimestre": trimestre,
            "tabla": tabla,
            "rango_inicio": start_row,
            "rango_fin": end_row
        })

    return blocks


# =========================================================
# PDF
# =========================================================
def build_pdf_all_lines(data_lineas, delegacion_general):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=25,
        leftMargin=25,
        topMargin=30,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "title_custom",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=15,
        leading=18,
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        "normal_custom",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        spaceAfter=4
    )

    small_style = ParagraphStyle(
        "small_custom",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=9
    )

    elements = []
    elements.append(Paragraph("REPORTE TRIMESTRAL DE LÍNEAS DE ACCIÓN", title_style))
    elements.append(Paragraph(f"<b>Delegación:</b> {safe_str(delegacion_general)}", normal_style))
    elements.append(Spacer(1, 8))

    ordered_keys = list(data_lineas.keys())

    for key in ordered_keys:
        item = data_lineas[key]
        info = item["info"]
        df = item["tabla"]
        trimestre = item["trimestre"]
        display_linea = item.get("display_linea", key)

        elements.append(Paragraph(f"<b>Línea de acción #:</b> {safe_str(display_linea)}", normal_style))
        elements.append(Paragraph(f"<b>Problemática:</b> {safe_str(info.get('problematica', ''))}", normal_style))
        elements.append(Paragraph(f"<b>Líder Estratégico:</b> {safe_str(info.get('lider', ''))}", normal_style))
        elements.append(Paragraph(f"<b>Trimestre:</b> {safe_str(trimestre)}", normal_style))
        elements.append(Spacer(1, 6))

        table_data = [[
            "Indicador",
            "Meta",
            "Avance",
            "Descripción",
            "Cantidad",
            "Observaciones"
        ]]

        for _, row in df.iterrows():
            table_data.append([
                Paragraph(safe_str(row.get("Indicador", "")), small_style),
                Paragraph(safe_str(row.get("Meta (editable)", "")), small_style),
                Paragraph(safe_str(row.get("Avance", "")), small_style),
                Paragraph(safe_str(row.get("Descripción", "")), small_style),
                Paragraph(safe_str(row.get("Cantidad", "")), small_style),
                Paragraph(safe_str(row.get("Observaciones (Editable)", "")), small_style),
            ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[120, 80, 55, 90, 55, 120]
        )

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 12))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# =========================================================
# ESTILOS
# =========================================================
st.markdown("""
<style>
.block-container {
    max-width: 1500px;
    padding-top: 1rem;
    padding-bottom: 2rem;
}

.form-box {
    border: 1px solid #444;
    padding: 18px;
    margin-top: 10px;
    margin-bottom: 16px;
    border-radius: 10px;
}

.line-card {
    border: 1px solid #333;
    border-radius: 10px;
    padding: 16px;
    margin-bottom: 18px;
}

.small-note {
    font-size: 0.92rem;
    opacity: 0.9;
}
</style>
""", unsafe_allow_html=True)


# =========================================================
# INTERFAZ
# =========================================================
st.title("Seguimiento de líneas de acción")

uploaded_file = st.file_uploader(
    "Arrastra y suelta el archivo .xlsm o .xlsx",
    type=["xlsm", "xlsx"]
)

if uploaded_file is not None:
    try:
        if st.session_state["archivo_nombre"] != uploaded_file.name:
            st.session_state["archivo_nombre"] = uploaded_file.name
            st.session_state["lineas_guardadas"] = {}
            st.session_state["pdf_final"] = None

        wb = load_workbook(
            BytesIO(uploaded_file.read()),
            data_only=False,
            keep_vba=True
        )

        main_sheet = find_best_main_sheet(wb)
        ws = wb[main_sheet]
        blocks = extract_blocks_from_sheet(ws)
        delegacion = get_delegacion(ws)

        if not blocks:
            st.warning("No se encontraron líneas de acción en la hoja.")
            st.stop()

        st.success(f"Hoja detectada: {main_sheet}")
        st.info(f"Líneas detectadas: {len(blocks)}")

        st.markdown('<div class="form-box">', unsafe_allow_html=True)
        c1, c2 = st.columns([1.2, 5])
        with c1:
            st.markdown("### Delegación :")
        with c2:
            st.text_input(
                "Delegación",
                value=delegacion,
                disabled=True,
                label_visibility="collapsed"
            )
        st.markdown("</div>", unsafe_allow_html=True)

        for idx, bloque in enumerate(blocks):
            linea_id = str(bloque["linea_accion"]).strip() if bloque["linea_accion"] else str(idx + 1)

            block_key = f"bloque_{idx}_{linea_id}"
            save_key = f"{idx}_{linea_id}"

            if save_key in st.session_state["lineas_guardadas"]:
                df_base = st.session_state["lineas_guardadas"][save_key]["tabla"].copy()
                trim_base = st.session_state["lineas_guardadas"][save_key]["trimestre"]
            else:
                df_base = bloque["tabla"].copy()
                trim_base = bloque["trimestre"] if bloque["trimestre"] in ["", "I", "II", "III", "IV"] else ""

            st.markdown('<div class="line-card">', unsafe_allow_html=True)
            st.subheader(f"Línea {linea_id}")

            c1, c2, c3, c4, c5, c6 = st.columns([1.8, 1.2, 1.6, 3.4, 1.8, 2.0])

            with c1:
                st.markdown("### línea de acción #:")
            with c2:
                st.text_input(
                    f"Línea de acción {block_key}",
                    value=linea_id,
                    disabled=True,
                    label_visibility="collapsed",
                    key=f"linea_{block_key}"
                )

            with c3:
                st.markdown("### Problemática:")
            with c4:
                st.text_input(
                    f"Problemática {block_key}",
                    value=bloque["problematica"],
                    disabled=True,
                    label_visibility="collapsed",
                    key=f"problematica_{block_key}"
                )

            with c5:
                st.markdown("### Líder Estratégico:")
            with c6:
                st.text_input(
                    f"Líder {block_key}",
                    value=bloque["lider"],
                    disabled=True,
                    label_visibility="collapsed",
                    key=f"lider_{block_key}"
                )

            c7, c8, c9 = st.columns([1.4, 1.2, 4])
            with c7:
                st.markdown("### Trimestre:")
            with c8:
                trim_options = ["", "I", "II", "III", "IV"]
                selected_trim = st.selectbox(
                    f"Trimestre {block_key}",
                    trim_options,
                    index=trim_options.index(trim_base if trim_base in trim_options else ""),
                    label_visibility="collapsed",
                    key=f"trim_{block_key}"
                )

            st.markdown("#### Detalle")

            # Avance, Descripción y Cantidad ya vienen del libro y quedan editables
            df_editado = st.data_editor(
                df_base,
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key=f"tabla_{block_key}"
            )

            c_btn1, c_btn2 = st.columns([2.2, 5])

            with c_btn1:
                if st.button(f"Guardar / Actualizar línea {linea_id}", key=f"guardar_{block_key}"):
                    st.session_state["lineas_guardadas"][save_key] = {
                        "display_linea": linea_id,
                        "info": {
                            "delegacion": delegacion,
                            "linea_accion": linea_id,
                            "problematica": bloque["problematica"],
                            "lider": bloque["lider"],
                            "rango_inicio": bloque["rango_inicio"],
                            "rango_fin": bloque["rango_fin"],
                        },
                        "tabla": df_editado.copy(),
                        "trimestre": selected_trim
                    }
                    st.session_state["pdf_final"] = None
                    st.success(f"Línea {linea_id} guardada correctamente")

            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("## Líneas guardadas")
        if st.session_state["lineas_guardadas"]:
            cols_saved = st.columns(4)
            ordered_items = list(st.session_state["lineas_guardadas"].items())

            for i, (_, item) in enumerate(ordered_items):
                with cols_saved[i % 4]:
                    st.write(f"✔ Línea {item.get('display_linea', '')}")
        else:
            st.info("Todavía no has guardado ninguna línea.")

        st.markdown("## Reporte final")

        if st.button("Preparar PDF con todas las líneas guardadas"):
            if not st.session_state["lineas_guardadas"]:
                st.warning("Primero debes guardar al menos una línea.")
            else:
                pdf_bytes = build_pdf_all_lines(
                    st.session_state["lineas_guardadas"],
                    delegacion_general=delegacion
                )
                st.session_state["pdf_final"] = pdf_bytes
                st.success("PDF generado correctamente.")

        if st.session_state["pdf_final"]:
            st.download_button(
                "Descargar PDF completo",
                data=st.session_state["pdf_final"],
                file_name=f"reporte_trimestral_{delegacion or 'delegacion'}.pdf",
                mime="application/pdf"
            )

        with st.expander("Resumen técnico"):
            debug_rows = []
            for b in blocks:
                debug_rows.append({
                    "Línea": b["linea_accion"],
                    "Problemática": b["problematica"],
                    "Líder": b["lider"],
                    "Trimestre detectado": b["trimestre"],
                    "Filas detalle": len(b["tabla"]),
                    "Fila inicio": b["rango_inicio"],
                    "Fila fin": b["rango_fin"],
                })
            st.dataframe(pd.DataFrame(debug_rows), use_container_width=True, hide_index=True)

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")

else:
    st.info("Sube un archivo para comenzar.")
